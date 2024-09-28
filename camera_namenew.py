import requests
from openpyxl import Workbook, load_workbook

def get_webpage_size(url):
    try:
        response = requests.get(url)
        if response.status_code == 200:
            return len(response.content), 'Success'
        else:
            return None, 'Error: ' + str(response.status_code)
    except Exception as e:
        return None, 'Error: ' + str(e)

def main():
    # Load existing Excel file
    wb = load_workbook("cameraport.xlsx")
    ws = wb.active

    # Create a new Excel file for results
    wb_result = Workbook()
    ws_result = wb_result.active
    ws_result.append(['URL', 'Webpage Size', 'Status'])

    # Iterate through rows
    for i, row in enumerate(ws.iter_rows(min_row=2, max_col=3, values_only=True), start=2):
        url, col_b, col_c = row
        if url:
            print(f"Processing row {i}: {url}")
            size, status = get_webpage_size(url)
            if size is not None:
                ws_result.append([url, size, status])
            else:
                print(f"Failed to process URL in column A: {url}")
                if col_b:
                    print(f"Checking URL in column B: {col_b}")
                    size, status = get_webpage_size(col_b)
                    if size is not None:
                        ws_result.append([col_b, size, status])
                    else:
                        print(f"Failed to process URL in column B: {col_b}")
                        if col_c:
                            print(f"Checking URL in column C: {col_c}")
                            size, status = get_webpage_size(col_c)
                            if size is not None:
                                ws_result.append([col_c, size, status])
                            else:
                                print(f"Failed to process URL in column C: {col_c}")
                                ws_result.append(['', '', 'All URLs failed'])
                        else:
                            print("No URL found in column C")
                            ws_result.append(['', '', 'Column C URL failed'])
                else:
                    print("No URL found in column B")
                    ws_result.append(['', '', 'Column B URL failed'])

    # Save the result to a new Excel file
    wb_result.save("Camera_name_results.xlsx")

if __name__ == "__main__":
    main()
