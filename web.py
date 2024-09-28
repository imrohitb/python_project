import requests
from bs4 import BeautifulSoup
import openpyxl
import os
import keyboard
import socket
import time
from urllib.parse import urlparse  # Import the urlparse function

# Rest of the code...

# Function to check if a port is open
def is_port_open(host, port):
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.settimeout(5)  # Adjust the timeout as needed
    result = sock.connect_ex((host, port))
    sock.close()
    return result == 0

# Function to get the page size in bytes
def get_page_size_bytes(url):
    try:
        # Extract host and port from the URL
        parsed_url = urlparse(url)
        host = parsed_url.netloc
        port = 80  # Default HTTP port

        # Check if a custom port is specified in the URL
        if ':' in host:
            host, port = host.split(':')
            port = int(port)

        if is_port_open(host, port):
            response = requests.get(url, timeout=5)
            if response.status_code == 200:
                content_length = len(response.content)  # Size in bytes
                return url, content_length, "Success"
            else:
                return url, None, f"Failed with status code {response.status_code}"
        else:
            return url, None, f"Port {port} is not open"

    except Exception as e:
        return url, None, str(e)

# Function to save the URL and page size results to an Excel file
def save_to_excel(results, excel_filename):
    if not os.path.exists(excel_filename):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(["URL", "Page Size (Bytes)", "Status"])
    else:
        workbook = openpyxl.load_workbook(excel_filename)
        worksheet = workbook.active

    for url, page_size, status in results:
        worksheet.append([url, page_size, status])

    workbook.save(excel_filename)

# Function to read URLs from an Excel file in column A
def read_urls_from_excel(filename):
    try:
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook.active
        urls = [cell.value for cell in worksheet['A'] if cell.value is not None]
        return urls
    except Exception as e:
        print(f"Error reading Excel file: {str(e)}")
        return []

if __name__ == "__main__":
    excel_filename = "results.xlsx"  # Specify your Excel file name here
    input_excel_filename = "urls.xlsx"  # Specify your input Excel file name here

    urls = read_urls_from_excel(input_excel_filename)
    paused = False  # Flag to track if the script is paused

    if not urls:
        print(f"No URLs found in '{input_excel_filename}' column A. Please add URLs to the Excel file.")
    else:
        results = []
        try:
            for url in urls:
                if not paused:
                    print(f"Checking port and processing URL: {url}")
                    try:
                        result = get_page_size_bytes(url)
                        results.append(result)
                        print(f"Status for URL '{url}': {result[2]}")
                    except requests.Timeout:
                        print(f"Timeout for URL '{url}'")
                else:
                    print("Script paused. Press Ctrl+Alt to resume...")

                # Check for key press (Ctrl+Alt to resume)
                if keyboard.is_pressed('ctrl+alt'):
                    paused = False
                    print("Script resumed.")
                    keyboard.press_and_release('ctrl+alt')  # Reset Ctrl+Alt status

                # Check for key press (Ctrl+C to pause)
                if keyboard.is_pressed('ctrl+c'):
                    paused = not paused  # Toggle pause state
                    print("Script paused. Press Ctrl+C again to exit or Ctrl+Alt to resume...")

        except KeyboardInterrupt:
            print("\nScript paused. Press Ctrl+C again to exit or Ctrl+Alt to resume...")

        save_to_excel(results, excel_filename)
        print(f"Results have been saved to '{excel_filename}'")
