import requests
from bs4 import BeautifulSoup
import openpyxl
import os
import keyboard
import socket
from urllib.parse import urlparse

# Function to check if a port is open
def is_port_open(host, port):
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.settimeout(5)  # Adjust the timeout as needed
    result = sock.connect_ex((host, port))
    sock.close()
    return result == 0

# Function to get the page size in bytes
def get_page_size_bytes(url, retry_ports=None):
    try:
        # Extract host and port from the URL
        parsed_url = urlparse(url)
        host = parsed_url.hostname
        default_port = parsed_url.port or 80  # Default HTTP port if not specified

        ports_to_try = retry_ports or [default_port]

        results = []
        for port in ports_to_try:
            if is_port_open(host, port):
                response = requests.get(f"{parsed_url.scheme}://{host}:{port}{parsed_url.path}", timeout=5)
                if response.status_code == 200:
                    content_length = len(response.content)  # Size in bytes
                    results.append((url, content_length, "Success", port, "Port Open"))
                    print(f"Port {port} is open for URL '{url}'")
                else:
                    results.append((url, None, f"Failed with status code {response.status_code}", port, "Port Open"))
                    print(f"Port {port} is open but failed for URL '{url}' with status code {response.status_code}")
            else:
                results.append((url, None, f"Port {port} is closed", port, "Port Closed"))
                print(f"Port {port} is closed for URL '{url}'")

        return results

    except Exception as e:
        return [(url, None, str(e), None, "Error")]

# Function to save the URL, page size, port status, and page status to an Excel file
def save_to_excel(results, excel_filename):
    if not os.path.exists(excel_filename):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(["URL", "Page Size (Bytes)", "Page Status", "Port", "Port Status"])
    else:
        workbook = openpyxl.load_workbook(excel_filename)
        worksheet = workbook.active

    for url, content_length, page_status, port, port_status in results:
        worksheet.append([url, content_length, page_status, port, port_status])

    workbook.save(excel_filename)

# Function to read URLs from an Excel file in multiple columns (A to F)
def read_urls_from_excel(filename):
    try:
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook.active
        urls = []

        for col in range(1, 7):  # Columns A to F
            column_letter = chr(ord('A') + col - 1)
            urls += [cell.value for cell in worksheet[column_letter] if cell.value is not None]

        return urls
    except Exception as e:
        print(f"Error reading Excel file: {str(e)}")
        return []

if __name__ == "__main__":
    excel_filename = "results.xlsx"  # Specify your Excel file name here
    input_excel_filename = "urls.xlsx"  # Specify your input Excel file name here

    urls = read_urls_from_excel(input_excel_filename)
    paused = False  # Flag to track if the script is paused

    if not urls:
        print(f"No URLs found in '{input_excel_filename}' columns A to F. Please add URLs to the Excel file.")
    else:
        results = []
        try:
            for url in urls:
                if not paused:
                    print(f"Checking ports and processing URL: {url}")
                    try:
                        result = get_page_size_bytes(url, retry_ports=[83, 84, 86, 1025, 1026, 1027])
                        results.extend(result)
                        print(f"Status for URL '{url}': {result}")
                    except requests.Timeout:
                        print(f"Timeout for URL '{url}'")
                else:
                    print("Script paused. Press Ctrl+Alt to resume...")

                # Check for key press (Ctrl+Alt to resume)
                if keyboard.is_pressed('ctrl+alt'):
                    paused = False
                    print("Script resumed.")
                    keyboard.press_and_release('ctrl+alt')  # Reset Ctrl+Alt status

                # Check for key press (Ctrl+C to pause)
                if keyboard.is_pressed('ctrl+c'):
                    paused = not paused  # Toggle pause state
                    print("Script paused. Press Ctrl+C again to exit or Ctrl+Alt to resume...")

        except KeyboardInterrupt:
            print("\nScript paused. Press Ctrl+C again to exit or Ctrl+Alt to resume...")

        save_to_excel(results, excel_filename)
        print(f"Results have been saved to '{excel_filename}'")
