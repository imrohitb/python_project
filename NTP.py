import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import requests
from requests.auth import HTTPDigestAuth
from openpyxl import Workbook
import threading

# Function to test connectivity
def test_connectivity(ip, username, password):
    url = f"http://{ip}:81/cgi-bin/magicBox.cgi?action=getSystemInfo"
    try:
        response = requests.get(url, auth=HTTPDigestAuth(username, password))
        if response.status_code == 200:
            return True, response.text
        else:
            return False, f"Failed to connect, HTTP Status Code: {response.status_code}\nResponse: {response.text}"
    except requests.exceptions.RequestException as e:
        return False, f"An error occurred: {e}"

# Function to update NTP settings
def update_ntp(ip, username, password, port, ntp_address):
    url = f"http://{ip}:{port}/cgi-bin/configManager.cgi?action=setConfig&NTP.Address={ntp_address}&NTP.Enable=true"
    response = requests.get(url, auth=HTTPDigestAuth(username, password))
    return response.status_code, response.text

# Function to load the Excel file
def load_excel_file():
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    if filepath:
        excel_path.set(filepath)

# Function to run the NTP update process
def run_ntp_update():
    if not excel_path.get():
        messagebox.showerror("Error", "Please select an Excel file.")
        return

    ntp_address = '192.168.100.24'
    workbook = openpyxl.load_workbook(excel_path.get())
    sheet = workbook.active

    result_workbook = Workbook()
    result_sheet = result_workbook.active
    result_sheet.append(["IP Address", "Connectivity Status", "NTP Update Status", "Response"])

    log_text.set("Starting NTP update process...\n")

    for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming the first row is the header
        ip, username, password, port = row
        log_text.set(log_text.get() + f"Testing connectivity for IP: {ip}\n")
        connected, message = test_connectivity(ip, username, password)
        if connected:
            log_text.set(log_text.get() + f"Successfully connected to {ip}. Updating NTP settings...\n")
            status_code, response_text = update_ntp(ip, username, password, port, ntp_address)
            ntp_status = f"Status: {status_code}"
            log_text.set(log_text.get() + f"IP: {ip} - Status: {status_code} - Response: {response_text}\n")
        else:
            ntp_status = "N/A"
            response_text = message
            log_text.set(log_text.get() + f"Failed to connect to {ip}. Message: {message}\n")

        result_sheet.append([ip, "Connected" if connected else "Failed", ntp_status, response_text])

    result_workbook.save("ntp_update_results.xlsx")
    log_text.set(log_text.get() + "NTP update process completed. Results saved to 'ntp_update_results.xlsx'.\n")
    messagebox.showinfo("Info", "NTP update process completed. Results saved to 'ntp_update_results.xlsx'.")

# Function to start the process in a separate thread
def start_process():
    threading.Thread(target=run_ntp_update).start()

# Create the main window
root = tk.Tk()
root.title("NTP Update Application")

excel_path = tk.StringVar()
log_text = tk.StringVar()

# Create and place widgets
tk.Label(root, text="Select Excel File:").grid(row=0, column=0, padx=10, pady=10)
tk.Entry(root, textvariable=excel_path, width=50).grid(row=0, column=1, padx=10, pady=10)
tk.Button(root, text="Browse", command=load_excel_file).grid(row=0, column=2, padx=10, pady=10)

tk.Button(root, text="Start", command=start_process).grid(row=1, column=1, padx=10, pady=10)

tk.Label(root, text="Log:").grid(row=2, column=0, padx=10, pady=10)
log_text_area = tk.Text(root, width=80, height=20)
log_text_area.grid(row=3, column=0, columnspan=3, padx=10, pady=10)

# Update the log area in real-time
def update_log():
    log_text_area.delete(1.0, tk.END)
    log_text_area.insert(tk.END, log_text.get())
    root.after(1000, update_log)

update_log()

# Run the main loop
root.mainloop()
